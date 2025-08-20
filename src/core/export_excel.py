# src/core/export_excel.py
from __future__ import annotations

from collections import OrderedDict, defaultdict
from typing import Any, Dict, Iterable, List, Tuple, Union
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.exceptions import IllegalCharacterError
import json
import re

JsonDict = Dict[str, Any]

# ------------------------------------------------------------
# Saneo para Excel (evita IllegalCharacterError)
# ------------------------------------------------------------
_ILLEGAL_RX = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")

def _excel_clean(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (int, float, bool)):
        return v
    s = str(v)
    # normaliza saltos de línea y quita no imprimibles
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = _ILLEGAL_RX.sub("", s)
    # Excel no admite cadenas muy largas en ciertas versiones; corta con indicación
    if len(s) > 32000:
        s = s[:31980] + " …(truncado)"
    return s


# ------------------------------------------------------------
# JSON helpers
# ------------------------------------------------------------
def _find_largest_json_object(text: str) -> str | None:
    """Devuelve el mayor bloque '{...}' balanceado en el texto, si existe."""
    start_idx = []
    best = None
    best_len = 0
    for i, ch in enumerate(text):
        if ch == "{":
            start_idx.append(i)
        elif ch == "}" and start_idx:
            s = start_idx.pop()
            length = i - s + 1
            if length > best_len:
                best = text[s : i + 1]
                best_len = length
    return best


def _try_parse_json_from_text(s: str) -> Any | None:
    """Intenta parsear JSON desde una cadena que puede contener basura alrededor."""
    if not s or not isinstance(s, str):
        return None

    # 1) intento directo
    try:
        return json.loads(s)
    except Exception:
        pass

    # 2) si parece array WAMP, intenta sacar el primer objeto
    #    formato típico: [16, 11, {}, "Topic", [{ ...obj... }]]
    try:
        arr = json.loads(s.strip())
        if isinstance(arr, list):
            # recorre buscando el primer dict
            stack = list(arr)
            while stack:
                it = stack.pop(0)
                if isinstance(it, dict):
                    return it
                if isinstance(it, list):
                    stack = it + stack
    except Exception:
        pass

    # 3) extrae el mayor bloque {...} y parsea
    candidate = _find_largest_json_object(s)
    if candidate:
        try:
            return json.loads(candidate)
        except Exception:
            return None
    return None


def _first_dict_in_nested(value: Any) -> JsonDict | None:
    """Devuelve el primer dict que encuentre dentro de una estructura (list/tuple/dict)."""
    if isinstance(value, dict):
        return value
    if isinstance(value, (list, tuple)):
        for it in value:
            d = _first_dict_in_nested(it)
            if d is not None:
                return d
    return None


CAND_TEXT_KEYS = (
    "json",
    "payload",
    "payload_text",
    "raw_text",
    "text",
    "data",
    "message",
    "body",
    "content",
)

RAW_KEYS = ("raw", "raw_text", "raw_payload", "payload", "json")


def extract_json_object(rec: JsonDict) -> Tuple[JsonDict | None, str | None]:
    """
    Intenta obtener un objeto JSON 'útil' desde el registro:
      - Si hay un dict ya en 'json' u otras claves, úsalo.
      - Si hay texto con JSON, parsea.
      - Si es array WAMP, extrae el dict interior.
    Devuelve (obj_dict, raw_text_detected).
    """
    # 1) dict directo
    for k in ("json_obj", "json", "obj", "object", "payload_obj"):
        if isinstance(rec.get(k), dict):
            return rec[k], json.dumps(rec[k], ensure_ascii=False)

    # 2) texto con JSON
    for k in CAND_TEXT_KEYS:
        v = rec.get(k)
        if isinstance(v, str) and v:
            parsed = _try_parse_json_from_text(v)
            if isinstance(parsed, dict):
                return parsed, v
            # si parsed es lista WAMP, coge el primer dict
            d = _first_dict_in_nested(parsed)
            if isinstance(d, dict):
                return d, v

    # 3) a veces args/kwargs esconden el objeto
    args = rec.get("args")
    kwargs = rec.get("kwargs")
    if isinstance(kwargs, dict) and kwargs:
        return kwargs, json.dumps(kwargs, ensure_ascii=False)
    d = _first_dict_in_nested(args)
    if isinstance(d, dict):
        return d, json.dumps(args, ensure_ascii=False)

    # nada encontrado
    raw_example = None
    for k in RAW_KEYS:
        if isinstance(rec.get(k), str):
            raw_example = rec[k]
            break
    return None, raw_example


# ------------------------------------------------------------
# Flatten
# ------------------------------------------------------------
def _is_scalar(x: Any) -> bool:
    return isinstance(x, (int, float, str, bool)) or x is None


def flatten_json(obj: Any, prefix: str = "") -> "OrderedDict[str, Any]":
    """
    Aplana dict/list a par clave->valor.
    - dict -> 'prefix.key'
    - list escalar -> 'prefix[0]', 'prefix[1]'
    - list de dicts -> idem (indexada)
    """
    out: "OrderedDict[str, Any]" = OrderedDict()
    p = f"{prefix}." if prefix else ""

    if isinstance(obj, dict):
        for k in sorted(obj.keys()):
            out.update(flatten_json(obj[k], f"{p}{k}" if prefix else k))
        return out

    if isinstance(obj, list):
        # si todos son escalares, indexa
        for i, it in enumerate(obj):
            out.update(flatten_json(it, f"{prefix}[{i}]"))
        return out

    # escalar
    out[prefix] = obj
    return out


# ------------------------------------------------------------
# Cabeceras: agrupación y colores
# ------------------------------------------------------------
PALETTE = [
    "FF1F4E79",  # azul oscuro
    "FF2E75B6",  # azul
    "FF2372A3",  # teal
    "FF548235",  # verde
    "FF7F6000",  # marrón
    "FF9E480E",  # naranja
    "FF984807",  # amber
    "FF5F497A",  # morado
    "FF44546A",  # gris azulado
    "FF70AD47",  # verde claro
]

HEADER_FONT = Font(color="FFFFFFFF", bold=True)
HEADER_ALIGN = Alignment(vertical="center", wrap_text=True)

def _top_prefix(col: str) -> str:
    """Prefijo para agrupar por color (antes del primer '.')"""
    if "[" in col and "." not in col:
        # ej: array en raíz -> toma nombre base
        return col.split("[", 1)[0]
    return col.split(".", 1)[0] if "." in col else col


# ------------------------------------------------------------
# Export principal
# ------------------------------------------------------------
def export_to_xlsx(records: List[JsonDict], out_path: str) -> None:
    """
    Exporta a Excel con 3 hojas:
      - Mensajes: metadatos + JSON aplanado (sin args/kwargs)
      - Raw: registro bruto + args/kwargs + texto crudo detectado
      - Resumen: conteos por type/topic
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mensajes"

    ws_raw = wb.create_sheet("Raw")
    ws_sum = wb.create_sheet("Resumen")

    # --- Detecta columnas de metadatos presentes en los registros ---
    meta_candidates = [
        "time", "time_hms_ms", "hms", "ms",
        "epoch", "frame", "tcp_stream", "stream",
        "src", "dst", "src_ip", "dst_ip", "src_port", "dst_port",
        "realm", "topic", "type", "opcode", "len", "proto",
    ]
    meta_cols: List[str] = [k for k in meta_candidates if any(k in r for r in records)]

    # --- Recolecta claves aplanadas para cabecera ---
    flat_keys_order: List[str] = []
    flat_keys_seen = set()

    extracted_objs: List[Tuple[JsonDict | None, str | None]] = []
    for rec in records:
        obj, rawtxt = extract_json_object(rec)
        extracted_objs.append((obj, rawtxt))
        if isinstance(obj, dict):
            flat = flatten_json(obj)
            for k in flat.keys():
                if k not in flat_keys_seen:
                    flat_keys_seen.add(k)
                    flat_keys_order.append(k)

    # Cabecera: meta + flat
    header = meta_cols + flat_keys_order
    if not header:
        header = ["time", "type"]  # fallback

    # --- Escribe cabecera con colores por grupo ---
    group2color: Dict[str, str] = {}
    pal_idx = 0
    for col_idx, col in enumerate(header, start=1):
        grp = _top_prefix(col) if col not in meta_cols else "__meta__"
        if grp not in group2color:
            # meta en gris oscuro, resto en paleta
            if grp == "__meta__":
                group2color[grp] = "FF323232"
            else:
                group2color[grp] = PALETTE[pal_idx % len(PALETTE)]
                pal_idx += 1
        fill = PatternFill(start_color=group2color[grp], end_color=group2color[grp], fill_type="solid")
        cell = ws.cell(row=1, column=col_idx, value=_excel_clean(col))
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    # --- Filas de datos (Mensajes) ---
    for rec, (obj, rawtxt) in zip(records, extracted_objs):
        row: List[Any] = []
        # metadatos
        for m in meta_cols:
            row.append(_excel_clean(rec.get(m)))
        # JSON aplanado (sin args/kwargs)
        flat = flatten_json(obj) if isinstance(obj, dict) else OrderedDict()
        for k in flat_keys_order:
            v = flat.get(k, "")
            # convierte listas/dicts restantes a JSON compacto
            if not _is_scalar(v):
                try:
                    v = json.dumps(v, ensure_ascii=False, separators=(",", ":"))
                except Exception:
                    v = str(v)
            row.append(_excel_clean(v))
        try:
            ws.append(row)
        except IllegalCharacterError:
            ws.append([_excel_clean(x) for x in row])

    # Ajuste de anchuras (simple)
    for col in ws.columns:
        max_len = 8
        for cell in col[:200]:  # muestreo
            v = cell.value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = min(l, 60)
        ws.column_dimensions[col[0].column_letter].width = max_len + 1

    # --- Hoja Raw ---
    # Claves base a exportar sin perder info (incluye args/kwargs/raw)
    raw_keys = sorted(
        set().union(*(r.keys() for r in records))
        | {"raw_detected_json_text"}
    )
    # Cabecera Raw
    for j, k in enumerate(raw_keys, start=1):
        c = ws_raw.cell(row=1, column=j, value=_excel_clean(k))
        c.fill = PatternFill("solid", start_color="FF1F4E79", end_color="FF1F4E79")
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN

    # Filas Raw
    for rec, (obj, rawtxt) in zip(records, extracted_objs):
        row = []
        for k in raw_keys:
            if k == "raw_detected_json_text":
                row.append(_excel_clean(rawtxt))
            else:
                v = rec.get(k, "")
                if isinstance(v, (dict, list)):
                    try:
                        v = json.dumps(v, ensure_ascii=False)
                    except Exception:
                        v = str(v)
                row.append(_excel_clean(v))
        try:
            ws_raw.append(row)
        except IllegalCharacterError:
            ws_raw.append([_excel_clean(x) for x in row])

    # --- Hoja Resumen ---
    by_type = defaultdict(int)
    by_topic = defaultdict(int)
    for rec in records:
        if "type" in rec and rec["type"]:
            by_type[str(rec["type"])] += 1
        if "topic" in rec and rec["topic"]:
            by_topic[str(rec["topic"])] += 1

    ws_sum.append(["Métrica", "Valor"])
    ws_sum.append(["Total registros", len(records)])
    ws_sum.append([])
    ws_sum.append(["Por type", "count"])
    for k, v in sorted(by_type.items(), key=lambda x: (-x[1], x[0])):
        ws_sum.append([k, v])
    ws_sum.append([])
    ws_sum.append(["Por topic", "count"])
    for k, v in sorted(by_topic.items(), key=lambda x: (-x[1], x[0])):
        ws_sum.append([k, v])

    # Estilo sencillo en Resumen
    for cell in ws_sum["A1:B1"][0]:
        cell.fill = PatternFill("solid", start_color="FF1F4E79", end_color="FF1F4E79")
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    # --- Guarda ---
    wb.save(out_path)
